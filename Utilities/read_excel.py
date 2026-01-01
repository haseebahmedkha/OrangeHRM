# Import the openpyxl module to read/write Excel files (.xlsx)
from openpyxl import load_workbook


# ================================
# Function for DDT (Data-Driven Testing) login test data
# ================================
def get_login_data(filepath, sheetname):
    """
    Reads login data from an Excel sheet and returns it as a list of tuples.

    :param filepath: Path to the Excel file (.xlsx)
    :param sheetname: Name of the sheet containing login data
    :return: List of tuples (username, password, expected_result)
    """

    # Load the Excel workbook
    workbook = load_workbook(filepath)

    # Select the sheet by name
    sheet = workbook[sheetname]

    # Initialize an empty list to store data
    data = []

    # Iterate through all rows starting from row 2 (skip headers)
    # values_only=True returns cell values instead of Cell objects
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, expected_result = row  # Unpack row values
        data.append((username, password, expected_result))  # Append as tuple

    return data


# ================================
# Function to read Admin user test data
# ================================
def get_admin_user_data(filepath, sheetname):
    """
    Reads admin user data from an Excel sheet and returns first 7 columns of each row.

    :param filepath: Path to the Excel file (.xlsx)
    :param sheetname: Name of the sheet containing admin data
    :return: List of tuples containing first 7 columns of each row
    """

    # Load the workbook
    workbook = load_workbook(filepath)

    # Select the sheet by name
    sheet = workbook[sheetname]

    # Initialize a list to store admin data
    data = []

    # Iterate through all rows starting from row 2 (skip headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Take only the first 7 columns and append as a tuple
        data.append(row[:7])

    return data
